"""Structured data parser for Excel, CSV, and JSON files.

Extracts tabular data and discovers schema information.
"""

from __future__ import annotations

import csv
import io
import json
import logging
from pathlib import Path

from src.core.models import FragmentType
from src.evidence.parsers.base import BaseParser, ParsedFragment, ParseResult

logger = logging.getLogger(__name__)


class StructuredDataParser(BaseParser):
    """Parser for structured data formats: XLSX, CSV, JSON."""

    supported_formats = [".xlsx", ".xls", ".csv", ".json"]

    async def parse(self, file_path: str, file_name: str) -> ParseResult:
        """Parse a structured data file and extract fragments.

        Routes to the appropriate sub-parser based on file extension.
        """
        ext = Path(file_name).suffix.lower()
        try:
            if ext in (".xlsx", ".xls"):
                return await self._parse_excel(file_path)
            elif ext == ".csv":
                return await self._parse_csv(file_path)
            elif ext == ".json":
                return await self._parse_json(file_path)
            else:
                return ParseResult(error=f"Unsupported structured data format: {ext}")
        except Exception as e:
            logger.exception("Failed to parse structured data: %s", file_name)
            return ParseResult(error=f"Parse error: {e}")

    async def _parse_excel(self, file_path: str) -> ParseResult:
        """Extract data from Excel workbooks using openpyxl."""
        from openpyxl import load_workbook

        fragments: list[ParsedFragment] = []
        metadata: dict[str, str | int | float | bool | None] = {}

        wb = load_workbook(file_path, read_only=True, data_only=True)
        metadata["sheet_count"] = len(wb.sheetnames)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[str] = []
            headers: list[str] = []
            row_count = 0

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                cells = [str(cell) if cell is not None else "" for cell in row]
                if row_idx == 0:
                    headers = cells
                rows.append(" | ".join(cells))
                row_count += 1

            if rows:
                table_text = "\n".join(rows)
                fragments.append(
                    ParsedFragment(
                        fragment_type=FragmentType.TABLE,
                        content=table_text,
                        metadata={
                            "sheet_name": sheet_name,
                            "row_count": row_count,
                            "column_count": len(headers),
                        },
                    )
                )

                # Schema discovery: extract column names as entity
                if headers:
                    schema_info = json.dumps({"columns": headers, "sheet": sheet_name})
                    fragments.append(
                        ParsedFragment(
                            fragment_type=FragmentType.ENTITY,
                            content=schema_info,
                            metadata={"schema_type": "excel_schema", "sheet_name": sheet_name},
                        )
                    )

        wb.close()
        return ParseResult(fragments=fragments, metadata=metadata)  # type: ignore[arg-type]

    async def _parse_csv(self, file_path: str) -> ParseResult:
        """Extract data from CSV files."""
        fragments: list[ParsedFragment] = []
        metadata: dict[str, str | int | float | bool | None] = {}

        with open(file_path, encoding="utf-8", errors="replace", newline="") as f:
            content = f.read()

        reader = csv.reader(io.StringIO(content))
        rows: list[str] = []
        headers: list[str] = []
        row_count = 0

        for row_idx, row in enumerate(reader):
            cells = [str(cell) for cell in row]
            if row_idx == 0:
                headers = cells
            rows.append(" | ".join(cells))
            row_count += 1

        if rows:
            table_text = "\n".join(rows)
            fragments.append(
                ParsedFragment(
                    fragment_type=FragmentType.TABLE,
                    content=table_text,
                    metadata={"row_count": row_count, "column_count": len(headers)},
                )
            )

            # Schema discovery
            if headers:
                schema_info = json.dumps({"columns": headers})
                fragments.append(
                    ParsedFragment(
                        fragment_type=FragmentType.ENTITY,
                        content=schema_info,
                        metadata={"schema_type": "csv_schema"},
                    )
                )

        metadata["row_count"] = row_count
        return ParseResult(fragments=fragments, metadata=metadata)  # type: ignore[arg-type]

    async def _parse_json(self, file_path: str) -> ParseResult:
        """Extract data from JSON files."""
        fragments: list[ParsedFragment] = []
        metadata: dict[str, str | int | float | bool | None] = {}

        with open(file_path, encoding="utf-8") as f:
            data = json.load(f)

        if isinstance(data, list):
            metadata["record_count"] = len(data)
            # Convert list of objects to a table-like format
            content = json.dumps(data, indent=2, default=str)
            fragments.append(
                ParsedFragment(
                    fragment_type=FragmentType.TABLE,
                    content=content,
                    metadata={"record_count": len(data)},
                )
            )

            # Schema from first record
            if data and isinstance(data[0], dict):
                schema_info = json.dumps({"fields": list(data[0].keys())})
                fragments.append(
                    ParsedFragment(
                        fragment_type=FragmentType.ENTITY,
                        content=schema_info,
                        metadata={"schema_type": "json_schema"},
                    )
                )

        elif isinstance(data, dict):
            metadata["key_count"] = len(data)
            content = json.dumps(data, indent=2, default=str)
            fragments.append(
                ParsedFragment(
                    fragment_type=FragmentType.TEXT,
                    content=content,
                    metadata={"key_count": len(data)},
                )
            )

            # Schema: top-level keys
            schema_info = json.dumps({"fields": list(data.keys())})
            fragments.append(
                ParsedFragment(
                    fragment_type=FragmentType.ENTITY,
                    content=schema_info,
                    metadata={"schema_type": "json_schema"},
                )
            )
        else:
            fragments.append(
                ParsedFragment(
                    fragment_type=FragmentType.TEXT,
                    content=str(data),
                    metadata={},
                )
            )

        return ParseResult(fragments=fragments, metadata=metadata)  # type: ignore[arg-type]
