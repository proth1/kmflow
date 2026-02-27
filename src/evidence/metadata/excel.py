"""Excel metadata extractor.

Extracts sheet count, per-sheet row counts, column names,
and inferred data types from Excel workbooks.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from src.evidence.metadata.base import ExtractedMetadata, MetadataExtractor, clean_string

logger = logging.getLogger(__name__)


class ExcelMetadataExtractor(MetadataExtractor):
    """Extract metadata from Excel (.xlsx, .xls) files."""

    supported_extensions = [".xlsx", ".xls"]

    def extract(self, file_path: str, file_size_bytes: int | None = None) -> ExtractedMetadata:
        """Extract metadata from an Excel workbook.

        Returns sheet_count and per-sheet tabular_metadata including
        row_count, column_names, and inferred data_types.
        """
        from openpyxl import load_workbook

        path = Path(file_path)
        size = file_size_bytes if file_size_bytes is not None else path.stat().st_size

        metadata = ExtractedMetadata(file_size_bytes=size)

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            metadata.sheet_count = len(wb.sheetnames)

            # Extract workbook-level properties
            if wb.properties:
                metadata.title = clean_string(getattr(wb.properties, "title", None))
                metadata.author = clean_string(getattr(wb.properties, "creator", None))
                if wb.properties.created:
                    metadata.creation_date = wb.properties.created.isoformat()
                if wb.properties.modified:
                    metadata.modification_date = wb.properties.modified.isoformat()

            tabular: list[dict[str, Any]] = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_meta = _extract_sheet_metadata(ws, sheet_name)
                tabular.append(sheet_meta)

            metadata.tabular_metadata = tabular
            wb.close()
        except Exception:  # Intentionally broad: Excel files can be corrupt
            logger.warning("Failed to extract Excel metadata from %s", file_path)

        return metadata


def _extract_sheet_metadata(ws: Any, sheet_name: str) -> dict[str, Any]:
    """Extract metadata from a single worksheet.

    Returns dict with sheet_name, row_count, column_names, and data_types.
    """
    column_names: list[str] = []
    data_types: list[str] = []
    row_count = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            column_names = [str(cell) if cell is not None else "" for cell in row]
        else:
            # Infer types from first data row
            if row_idx == 1:
                data_types = [_infer_type(cell) for cell in row]
        row_count += 1

    return {
        "sheet_name": sheet_name,
        "row_count": row_count,
        "column_names": column_names,
        "data_types": data_types,
    }


def _infer_type(value: Any) -> str:
    """Infer a simple type name from a cell value."""
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "float"
    from datetime import datetime

    if isinstance(value, datetime):
        return "datetime"
    return "string"
