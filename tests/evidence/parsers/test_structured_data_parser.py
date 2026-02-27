"""Tests for StructuredDataParser (Excel, CSV, JSON) — Story #296.

Covers BDD Scenario 2 (Excel multi-sheet parsing) and Scenario 5 (corrupted files).
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from src.core.models import FragmentType
from src.evidence.parsers.structured_data_parser import StructuredDataParser


@pytest.fixture
def parser() -> StructuredDataParser:
    return StructuredDataParser()


# ---------------------------------------------------------------------------
# Supported formats
# ---------------------------------------------------------------------------


class TestSupportedFormats:
    def test_supported_formats(self, parser: StructuredDataParser) -> None:
        for ext in [".xlsx", ".xls", ".csv", ".json"]:
            assert parser.can_parse(ext), f"{ext} should be supported"

    def test_unsupported(self, parser: StructuredDataParser) -> None:
        assert not parser.can_parse(".pdf")
        assert not parser.can_parse(".bpmn")


# ---------------------------------------------------------------------------
# BDD Scenario 2: Excel file with multiple sheets
# ---------------------------------------------------------------------------


class TestBDDScenario2ExcelParsing:
    """Scenario 2: Excel file with multiple worksheets parsed."""

    @pytest.mark.asyncio
    async def test_excel_multi_sheet_fragments(self, parser: StructuredDataParser, tmp_path: Path) -> None:
        """Each sheet produces a TABLE fragment."""
        xlsx_path = tmp_path / "multi.xlsx"
        _create_multi_sheet_excel(xlsx_path)

        result = await parser.parse(str(xlsx_path), "multi.xlsx")

        assert result.error is None
        table_frags = [f for f in result.fragments if f.fragment_type == FragmentType.TABLE]
        # At least 2 sheets → at least 2 table fragments
        assert len(table_frags) >= 2

    @pytest.mark.asyncio
    async def test_excel_sheet_count_metadata(self, parser: StructuredDataParser, tmp_path: Path) -> None:
        """Metadata includes sheet_count."""
        xlsx_path = tmp_path / "multi.xlsx"
        _create_multi_sheet_excel(xlsx_path)

        result = await parser.parse(str(xlsx_path), "multi.xlsx")

        assert result.metadata.get("sheet_count") == 2

    @pytest.mark.asyncio
    async def test_excel_row_count_in_fragment(self, parser: StructuredDataParser, tmp_path: Path) -> None:
        """Fragment metadata includes row_count."""
        xlsx_path = tmp_path / "data.xlsx"
        _create_single_sheet_excel(xlsx_path, rows=5)

        result = await parser.parse(str(xlsx_path), "data.xlsx")

        table_frags = [f for f in result.fragments if f.fragment_type == FragmentType.TABLE]
        assert len(table_frags) >= 1
        # 5 data rows + 1 header = 6
        assert table_frags[0].metadata.get("row_count") == 6

    @pytest.mark.asyncio
    async def test_excel_column_names_schema(self, parser: StructuredDataParser, tmp_path: Path) -> None:
        """Schema discovery extracts column names as ENTITY fragment."""
        xlsx_path = tmp_path / "data.xlsx"
        _create_single_sheet_excel(xlsx_path)

        result = await parser.parse(str(xlsx_path), "data.xlsx")

        entity_frags = [f for f in result.fragments if f.fragment_type == FragmentType.ENTITY]
        assert len(entity_frags) >= 1
        schema_data = json.loads(entity_frags[0].content)
        assert "columns" in schema_data
        assert "Name" in schema_data["columns"]


# ---------------------------------------------------------------------------
# CSV Parsing
# ---------------------------------------------------------------------------


class TestCsvParsing:
    @pytest.mark.asyncio
    async def test_csv_table_extraction(self, parser: StructuredDataParser, tmp_path: Path) -> None:
        csv_path = tmp_path / "data.csv"
        csv_path.write_text("Name,Age,City\nAlice,30,NYC\nBob,25,LA\n")

        result = await parser.parse(str(csv_path), "data.csv")

        assert result.error is None
        table_frags = [f for f in result.fragments if f.fragment_type == FragmentType.TABLE]
        assert len(table_frags) == 1
        assert "Alice" in table_frags[0].content

    @pytest.mark.asyncio
    async def test_csv_row_count(self, parser: StructuredDataParser, tmp_path: Path) -> None:
        csv_path = tmp_path / "data.csv"
        csv_path.write_text("A,B\n1,2\n3,4\n")

        result = await parser.parse(str(csv_path), "data.csv")

        assert result.metadata.get("row_count") == 3  # header + 2 data rows

    @pytest.mark.asyncio
    async def test_csv_schema_discovery(self, parser: StructuredDataParser, tmp_path: Path) -> None:
        csv_path = tmp_path / "data.csv"
        csv_path.write_text("Name,Age\nAlice,30\n")

        result = await parser.parse(str(csv_path), "data.csv")

        entity_frags = [f for f in result.fragments if f.fragment_type == FragmentType.ENTITY]
        assert len(entity_frags) >= 1
        schema = json.loads(entity_frags[0].content)
        assert "Name" in schema["columns"]
        assert "Age" in schema["columns"]


# ---------------------------------------------------------------------------
# JSON Parsing
# ---------------------------------------------------------------------------


class TestJsonParsing:
    @pytest.mark.asyncio
    async def test_json_array_parsing(self, parser: StructuredDataParser, tmp_path: Path) -> None:
        json_path = tmp_path / "data.json"
        data = [{"name": "Alice", "age": 30}, {"name": "Bob", "age": 25}]
        json_path.write_text(json.dumps(data))

        result = await parser.parse(str(json_path), "data.json")

        assert result.error is None
        assert result.metadata.get("record_count") == 2

    @pytest.mark.asyncio
    async def test_json_object_parsing(self, parser: StructuredDataParser, tmp_path: Path) -> None:
        json_path = tmp_path / "config.json"
        data = {"key": "value", "nested": {"a": 1}}
        json_path.write_text(json.dumps(data))

        result = await parser.parse(str(json_path), "config.json")

        assert result.error is None
        assert result.metadata.get("key_count") == 2

    @pytest.mark.asyncio
    async def test_json_schema_extraction(self, parser: StructuredDataParser, tmp_path: Path) -> None:
        json_path = tmp_path / "data.json"
        data = [{"id": 1, "name": "test"}]
        json_path.write_text(json.dumps(data))

        result = await parser.parse(str(json_path), "data.json")

        entity_frags = [f for f in result.fragments if f.fragment_type == FragmentType.ENTITY]
        assert len(entity_frags) >= 1
        schema = json.loads(entity_frags[0].content)
        assert "id" in schema["fields"]
        assert "name" in schema["fields"]


# ---------------------------------------------------------------------------
# BDD Scenario 5: Corrupted structured data
# ---------------------------------------------------------------------------


class TestCorruptedStructuredData:
    @pytest.mark.asyncio
    async def test_corrupted_excel(self, parser: StructuredDataParser, tmp_path: Path) -> None:
        bad_xlsx = tmp_path / "corrupt.xlsx"
        bad_xlsx.write_bytes(b"NOT EXCEL DATA")

        result = await parser.parse(str(bad_xlsx), "corrupt.xlsx")

        assert result.error is not None
        assert "Parse error" in result.error

    @pytest.mark.asyncio
    async def test_corrupted_json(self, parser: StructuredDataParser, tmp_path: Path) -> None:
        bad_json = tmp_path / "corrupt.json"
        bad_json.write_text("{invalid json content")

        result = await parser.parse(str(bad_json), "corrupt.json")

        assert result.error is not None
        assert "Parse error" in result.error

    @pytest.mark.asyncio
    async def test_unsupported_extension(self, parser: StructuredDataParser) -> None:
        result = await parser.parse("/tmp/fake.xyz", "fake.xyz")
        assert result.error is not None
        assert "Unsupported" in result.error


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _create_multi_sheet_excel(path: Path) -> None:
    """Create an Excel workbook with 2 sheets."""
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["Name", "Value"])
        ws1.append(["Item A", 100])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["ID", "Status"])
        ws2.append([1, "Active"])
        ws2.append([2, "Inactive"])

        wb.save(str(path))
    except ImportError:
        pytest.skip("openpyxl not available")


def _create_single_sheet_excel(path: Path, rows: int = 3) -> None:
    """Create a single-sheet Excel workbook."""
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Value", "Status"])
        for i in range(rows):
            ws.append([f"Item_{i}", i * 10, "active"])
        wb.save(str(path))
    except ImportError:
        pytest.skip("openpyxl not available")
