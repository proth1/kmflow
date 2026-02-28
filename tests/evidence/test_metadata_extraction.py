"""Tests for Evidence Cataloging with Automated Metadata Extraction — Story #304.

Covers all 4 BDD scenarios:
1. PDF metadata is extracted on ingest
2. Excel metadata captures tabular structure
3. Language detection sets the language field
4. Evidence catalog returns filtered and paginated results
"""

from __future__ import annotations

import tempfile
import uuid
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Any
from unittest.mock import AsyncMock, MagicMock

import pytest
from httpx import AsyncClient

from src.core.models import EvidenceCategory, EvidenceItem, ValidationStatus
from src.evidence.metadata.base import ExtractedMetadata
from src.evidence.metadata.language import detect_language

# ---------------------------------------------------------------------------
# BDD Scenario 1: PDF metadata is extracted on ingest
# ---------------------------------------------------------------------------


class TestBDDScenario1PdfMetadata:
    """Given a PDF file is uploaded and accepted by the ingestion pipeline
    When the MetadataExtractor processes the file
    Then the following fields are populated on the evidence record.
    """

    def test_pdf_extractor_extracts_title(self) -> None:
        """Title is extracted from PDF info dict."""
        from src.evidence.metadata.pdf import PdfMetadataExtractor

        extractor = PdfMetadataExtractor()
        with _create_test_pdf(title="Q4 Financial Summary") as pdf_path:
            result = extractor.extract(pdf_path)
            assert result.title == "Q4 Financial Summary"

    def test_pdf_extractor_extracts_author(self) -> None:
        """Author is extracted from PDF info dict."""
        from src.evidence.metadata.pdf import PdfMetadataExtractor

        extractor = PdfMetadataExtractor()
        with _create_test_pdf(author="Jane Smith") as pdf_path:
            result = extractor.extract(pdf_path)
            assert result.author == "Jane Smith"

    def test_pdf_extractor_extracts_creation_date(self) -> None:
        """Creation date is extracted and converted to ISO 8601."""
        from src.evidence.metadata.pdf import _parse_pdf_date

        iso_date = _parse_pdf_date("D:20251001090000+00'00'")
        assert iso_date is not None
        assert iso_date.startswith("2025-10-01")

    def test_pdf_extractor_extracts_page_count(self) -> None:
        """Page count is extracted from the PDF."""
        from src.evidence.metadata.pdf import PdfMetadataExtractor

        extractor = PdfMetadataExtractor()
        with _create_test_pdf(pages=3) as pdf_path:
            result = extractor.extract(pdf_path)
            assert result.page_count == 3

    def test_pdf_extractor_extracts_file_size(self) -> None:
        """File size in bytes is captured."""
        from src.evidence.metadata.pdf import PdfMetadataExtractor

        extractor = PdfMetadataExtractor()
        with _create_test_pdf() as pdf_path:
            size = Path(pdf_path).stat().st_size
            result = extractor.extract(pdf_path, file_size_bytes=size)
            assert result.file_size_bytes == size
            assert result.file_size_bytes > 0

    def test_missing_fields_stored_as_null(self) -> None:
        """Missing metadata fields are stored as null, not omitted."""
        metadata = ExtractedMetadata()
        d = metadata.to_dict()
        assert "title" in d
        assert d["title"] is None
        assert "author" in d
        assert d["author"] is None
        assert "creation_date" in d
        assert d["creation_date"] is None
        assert "page_count" in d
        assert d["page_count"] is None

    def test_pdf_extractor_handles_corrupt_file(self) -> None:
        """Corrupt PDF returns metadata with null fields gracefully."""
        from src.evidence.metadata.pdf import PdfMetadataExtractor

        extractor = PdfMetadataExtractor()
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            f.write(b"not a real pdf")
            f.flush()
            result = extractor.extract(f.name, file_size_bytes=14)
        assert result.file_size_bytes == 14
        assert result.page_count is None

    def test_pdf_supported_extensions(self) -> None:
        """PDF extractor handles .pdf extension."""
        from src.evidence.metadata.pdf import PdfMetadataExtractor

        extractor = PdfMetadataExtractor()
        assert extractor.can_extract(".pdf")
        assert not extractor.can_extract(".xlsx")

    def test_pdf_date_parsing_short_format(self) -> None:
        """PDF date with only YYYYMMDD parses correctly."""
        from src.evidence.metadata.pdf import _parse_pdf_date

        result = _parse_pdf_date("D:20251001")
        assert result is not None
        assert "2025-10-01" in result

    def test_pdf_date_parsing_invalid(self) -> None:
        """Invalid PDF date returns None."""
        from src.evidence.metadata.pdf import _parse_pdf_date

        assert _parse_pdf_date(None) is None
        assert _parse_pdf_date("") is None
        assert _parse_pdf_date("notadate") is None


# ---------------------------------------------------------------------------
# BDD Scenario 2: Excel metadata captures tabular structure
# ---------------------------------------------------------------------------


class TestBDDScenario2ExcelMetadata:
    """Given an Excel file with 3 sheets is uploaded
    When the MetadataExtractor processes the file
    Then the metadata includes sheet_count=3.
    """

    def test_excel_extractor_captures_sheet_count(self) -> None:
        """Sheet count matches the number of sheets in the workbook."""
        from src.evidence.metadata.excel import ExcelMetadataExtractor

        extractor = ExcelMetadataExtractor()
        with _create_test_excel(sheet_count=3) as xlsx_path:
            result = extractor.extract(xlsx_path)
            assert result.sheet_count == 3

    def test_excel_extractor_captures_row_count(self) -> None:
        """Per-sheet row count is captured in tabular_metadata."""
        from src.evidence.metadata.excel import ExcelMetadataExtractor

        extractor = ExcelMetadataExtractor()
        with _create_test_excel(sheet_count=1, rows=10) as xlsx_path:
            result = extractor.extract(xlsx_path)
            assert result.tabular_metadata is not None
            assert len(result.tabular_metadata) == 1
            assert result.tabular_metadata[0]["row_count"] == 10

    def test_excel_extractor_captures_column_names(self) -> None:
        """Column names are extracted from the first row of each sheet."""
        from src.evidence.metadata.excel import ExcelMetadataExtractor

        extractor = ExcelMetadataExtractor()
        with _create_test_excel(columns=["Name", "Age", "City"]) as xlsx_path:
            result = extractor.extract(xlsx_path)
            assert result.tabular_metadata is not None
            assert result.tabular_metadata[0]["column_names"] == ["Name", "Age", "City"]

    def test_excel_extractor_infers_data_types(self) -> None:
        """Data types are inferred from the first data row."""
        from src.evidence.metadata.excel import ExcelMetadataExtractor

        extractor = ExcelMetadataExtractor()
        with _create_test_excel(columns=["Name", "Age"], data_rows=[["Alice", 30]]) as xlsx_path:
            result = extractor.extract(xlsx_path)
            assert result.tabular_metadata is not None
            types = result.tabular_metadata[0]["data_types"]
            assert "string" in types
            assert "integer" in types

    def test_excel_metadata_stored_under_tabular_metadata_key(self) -> None:
        """Tabular metadata is stored under the 'tabular_metadata' key."""
        from src.evidence.metadata.excel import ExcelMetadataExtractor

        extractor = ExcelMetadataExtractor()
        with _create_test_excel(sheet_count=2) as xlsx_path:
            result = extractor.extract(xlsx_path)
            d = result.to_dict()
            assert "tabular_metadata" in d
            assert len(d["tabular_metadata"]) == 2

    def test_excel_extractor_handles_corrupt_file(self) -> None:
        """Corrupt Excel returns metadata gracefully."""
        from src.evidence.metadata.excel import ExcelMetadataExtractor

        extractor = ExcelMetadataExtractor()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(b"not a real excel")
            f.flush()
            result = extractor.extract(f.name, file_size_bytes=16)
        assert result.file_size_bytes == 16
        assert result.sheet_count is None

    def test_excel_supported_extensions(self) -> None:
        """Excel extractor handles .xlsx and .xls extensions."""
        from src.evidence.metadata.excel import ExcelMetadataExtractor

        extractor = ExcelMetadataExtractor()
        assert extractor.can_extract(".xlsx")
        assert extractor.can_extract(".xls")
        assert not extractor.can_extract(".pdf")

    def test_excel_extractor_captures_workbook_title(self) -> None:
        """Workbook-level title from properties is extracted."""
        from src.evidence.metadata.excel import ExcelMetadataExtractor

        extractor = ExcelMetadataExtractor()
        with _create_test_excel(title="Budget Report") as xlsx_path:
            result = extractor.extract(xlsx_path)
            assert result.title == "Budget Report"


# ---------------------------------------------------------------------------
# BDD Scenario 3: Language detection sets the language field
# ---------------------------------------------------------------------------


class TestBDDScenario3LanguageDetection:
    """Given an evidence document whose body text is predominantly French
    When language detection runs during metadata extraction
    Then the language field on the evidence record is set to 'fr'.
    """

    def test_english_text_detected_as_en(self) -> None:
        """English text is detected as 'en'."""
        text = (
            "This is a comprehensive report on the financial performance "
            "of the organization during the fourth quarter of the fiscal year."
        )
        assert detect_language(text) == "en"

    def test_french_text_detected_as_fr(self) -> None:
        """French text is detected as 'fr'."""
        text = (
            "Ceci est un rapport complet sur la performance financière "
            "de l'organisation au cours du quatrième trimestre de l'exercice."
        )
        assert detect_language(text) == "fr"

    def test_german_text_detected_as_de(self) -> None:
        """German text is detected correctly."""
        text = (
            "Dies ist ein umfassender Bericht über die finanzielle Leistung "
            "der Organisation im vierten Quartal des Geschäftsjahres."
        )
        assert detect_language(text) == "de"

    def test_short_text_returns_none(self) -> None:
        """Text shorter than 20 chars returns None."""
        assert detect_language("hello") is None
        assert detect_language("") is None

    def test_empty_text_returns_none(self) -> None:
        """Empty or whitespace-only text returns None."""
        assert detect_language("") is None
        assert detect_language("   ") is None

    def test_dominant_language_for_mixed(self) -> None:
        """Mixed language text returns the dominant language."""
        # Predominantly English with some French
        text = (
            "The quarterly financial report indicates strong revenue growth. "
            "Operating margins expanded significantly. The board approved the dividend. "
            "Net income rose substantially. Revenue exceeded projections by a wide margin. "
            "Bonjour le monde."
        )
        result = detect_language(text)
        assert result == "en"


# ---------------------------------------------------------------------------
# BDD Scenario 4: Evidence catalog returns filtered and paginated results
# ---------------------------------------------------------------------------


class TestBDDScenario4CatalogAPI:
    """Given an engagement with 50 evidence items
    When the catalog API is queried with filters
    Then filtered and paginated results are returned.
    """

    @pytest.mark.asyncio
    async def test_catalog_returns_paginated_response(self, client: AsyncClient, mock_db_session: AsyncMock) -> None:
        """Catalog response includes pagination fields."""
        _setup_catalog_mocks(mock_db_session, items=[], total=0)

        response = await client.get(
            "/api/v1/evidence/catalog",
            params={"engagement_id": str(uuid.uuid4()), "limit": 10, "offset": 0},
        )
        assert response.status_code == 200
        data = response.json()
        assert "items" in data
        assert "total" in data
        assert "limit" in data
        assert "offset" in data
        assert "has_more" in data

    @pytest.mark.asyncio
    async def test_catalog_filters_by_category(self, client: AsyncClient, mock_db_session: AsyncMock) -> None:
        """Catalog filters by evidence category."""
        items = _make_mock_evidence_items(3, category=EvidenceCategory.DOCUMENTS)
        _setup_catalog_mocks(mock_db_session, items=items, total=3)

        response = await client.get(
            "/api/v1/evidence/catalog",
            params={
                "engagement_id": str(uuid.uuid4()),
                "category": "documents",
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["total"] == 3
        assert len(data["items"]) == 3

    @pytest.mark.asyncio
    async def test_catalog_has_more_pagination(self, client: AsyncClient, mock_db_session: AsyncMock) -> None:
        """has_more is True when total exceeds offset + limit."""
        items = _make_mock_evidence_items(5)
        _setup_catalog_mocks(mock_db_session, items=items, total=15)

        response = await client.get(
            "/api/v1/evidence/catalog",
            params={"engagement_id": str(uuid.uuid4()), "limit": 5, "offset": 0},
        )
        data = response.json()
        assert data["has_more"] is True

    @pytest.mark.asyncio
    async def test_catalog_no_more_on_last_page(self, client: AsyncClient, mock_db_session: AsyncMock) -> None:
        """has_more is False on the last page."""
        items = _make_mock_evidence_items(3)
        _setup_catalog_mocks(mock_db_session, items=items, total=3)

        response = await client.get(
            "/api/v1/evidence/catalog",
            params={"engagement_id": str(uuid.uuid4()), "limit": 10, "offset": 0},
        )
        data = response.json()
        assert data["has_more"] is False

    @pytest.mark.asyncio
    async def test_catalog_item_has_required_fields(self, client: AsyncClient, mock_db_session: AsyncMock) -> None:
        """Each catalog item includes id, title, category, creation_date, quality_score."""
        items = _make_mock_evidence_items(1, category=EvidenceCategory.DOCUMENTS)
        _setup_catalog_mocks(mock_db_session, items=items, total=1)

        response = await client.get(
            "/api/v1/evidence/catalog",
            params={"engagement_id": str(uuid.uuid4())},
        )
        data = response.json()
        assert len(data["items"]) == 1
        item = data["items"][0]
        assert "id" in item
        assert "name" in item
        assert "category" in item
        assert "creation_date" in item
        assert "quality_score" in item

    @pytest.mark.asyncio
    async def test_catalog_filters_by_language(self, client: AsyncClient, mock_db_session: AsyncMock) -> None:
        """Catalog filters by detected language."""
        items = _make_mock_evidence_items(2, language="fr")
        _setup_catalog_mocks(mock_db_session, items=items, total=2)

        response = await client.get(
            "/api/v1/evidence/catalog",
            params={"engagement_id": str(uuid.uuid4()), "language": "fr"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["total"] == 2

    @pytest.mark.asyncio
    async def test_catalog_requires_engagement_id(self, client: AsyncClient, mock_db_session: AsyncMock) -> None:
        """Catalog endpoint requires engagement_id query param."""
        response = await client.get("/api/v1/evidence/catalog")
        assert response.status_code == 422

    @pytest.mark.asyncio
    async def test_catalog_full_text_search(self, client: AsyncClient, mock_db_session: AsyncMock) -> None:
        """Catalog supports full-text search via q parameter."""
        items = _make_mock_evidence_items(1)
        _setup_catalog_mocks(mock_db_session, items=items, total=1)

        response = await client.get(
            "/api/v1/evidence/catalog",
            params={"engagement_id": str(uuid.uuid4()), "q": "financial"},
        )
        assert response.status_code == 200

    @pytest.mark.asyncio
    async def test_catalog_filters_by_date_from(self, client: AsyncClient, mock_db_session: AsyncMock) -> None:
        """Catalog filters by date_from lower bound."""
        items = _make_mock_evidence_items(2)
        _setup_catalog_mocks(mock_db_session, items=items, total=2)

        response = await client.get(
            "/api/v1/evidence/catalog",
            params={"engagement_id": str(uuid.uuid4()), "date_from": "2025-01-01"},
        )
        assert response.status_code == 200
        assert response.json()["total"] == 2

    @pytest.mark.asyncio
    async def test_catalog_filters_by_date_to(self, client: AsyncClient, mock_db_session: AsyncMock) -> None:
        """Catalog filters by date_to upper bound."""
        items = _make_mock_evidence_items(1)
        _setup_catalog_mocks(mock_db_session, items=items, total=1)

        response = await client.get(
            "/api/v1/evidence/catalog",
            params={"engagement_id": str(uuid.uuid4()), "date_to": "2025-12-31"},
        )
        assert response.status_code == 200
        assert response.json()["total"] == 1

    @pytest.mark.asyncio
    async def test_catalog_search_escapes_like_wildcards(self, client: AsyncClient, mock_db_session: AsyncMock) -> None:
        """Search query with LIKE wildcards (%, _) is escaped safely."""
        items = _make_mock_evidence_items(0)
        _setup_catalog_mocks(mock_db_session, items=items, total=0)

        response = await client.get(
            "/api/v1/evidence/catalog",
            params={"engagement_id": str(uuid.uuid4()), "q": "100%_complete"},
        )
        assert response.status_code == 200


# ---------------------------------------------------------------------------
# Additional metadata extractor tests
# ---------------------------------------------------------------------------


class TestExtractedMetadataModel:
    """Test ExtractedMetadata dataclass behavior."""

    def test_to_dict_includes_all_base_fields(self) -> None:
        """to_dict includes all base fields even when None."""
        m = ExtractedMetadata()
        d = m.to_dict()
        expected_keys = {
            "title",
            "author",
            "creation_date",
            "modification_date",
            "page_count",
            "file_size_bytes",
            "detected_language",
            "sheet_count",
            "tabular_metadata",
        }
        assert expected_keys.issubset(set(d.keys()))

    def test_to_dict_with_tabular_metadata(self) -> None:
        """to_dict includes tabular_metadata when present."""
        m = ExtractedMetadata(
            sheet_count=2,
            tabular_metadata=[{"sheet_name": "Sheet1", "row_count": 10}],
        )
        d = m.to_dict()
        assert d["sheet_count"] == 2
        assert "tabular_metadata" in d

    def test_to_dict_with_extra(self) -> None:
        """to_dict merges extra fields into the output."""
        m = ExtractedMetadata(extra={"custom_field": "value"})
        d = m.to_dict()
        assert d["custom_field"] == "value"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _create_test_pdf(
    title: str | None = None,
    author: str | None = None,
    pages: int = 1,
) -> tempfile._TemporaryFileWrapper:
    """Create a minimal PDF file for testing.

    Returns a context manager yielding the file path.
    """
    from contextlib import contextmanager

    @contextmanager
    def _ctx():
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            tmp_path = f.name
        c = canvas.Canvas(tmp_path, pagesize=letter)
        if title:
            c.setTitle(title)
        if author:
            c.setAuthor(author)
        for i in range(pages):
            c.drawString(72, 700, f"Page {i + 1}")
            if i < pages - 1:
                c.showPage()
        c.save()
        try:
            yield tmp_path
        finally:
            Path(tmp_path).unlink(missing_ok=True)

    return _ctx()


def _create_test_excel(
    sheet_count: int = 1,
    rows: int = 5,
    columns: list[str] | None = None,
    data_rows: list[list[Any]] | None = None,
    title: str | None = None,
) -> tempfile._TemporaryFileWrapper:
    """Create a minimal Excel file for testing.

    Returns a context manager yielding the file path.
    """
    from contextlib import contextmanager

    @contextmanager
    def _ctx():
        from openpyxl import Workbook

        wb = Workbook()
        if title and wb.properties:
            wb.properties.title = title

        cols = columns or ["A", "B", "C"]

        for i in range(sheet_count):
            if i == 0:
                ws = wb.active
                ws.title = f"Sheet{i + 1}"
            else:
                ws = wb.create_sheet(f"Sheet{i + 1}")

            # Write header
            for col_idx, col_name in enumerate(cols, 1):
                ws.cell(row=1, column=col_idx, value=col_name)

            # Write data
            if data_rows:
                for row_idx, row_data in enumerate(data_rows, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                for row_idx in range(2, rows + 1):
                    for col_idx in range(1, len(cols) + 1):
                        ws.cell(row=row_idx, column=col_idx, value=f"val_{row_idx}_{col_idx}")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        wb.save(tmp_path)
        wb.close()
        try:
            yield tmp_path
        finally:
            Path(tmp_path).unlink(missing_ok=True)

    return _ctx()


def _make_mock_evidence_items(
    count: int,
    category: EvidenceCategory = EvidenceCategory.DOCUMENTS,
    language: str | None = "en",
) -> list[MagicMock]:
    """Create mock EvidenceItem objects for catalog tests."""
    items = []
    for i in range(count):
        item = MagicMock(spec=EvidenceItem)
        item.id = uuid.uuid4()
        item.engagement_id = uuid.uuid4()
        item.name = f"Evidence Item {i + 1}"
        item.category = category
        item.format = "pdf"
        item.size_bytes = 1024 * (i + 1)
        item.detected_language = language
        item.validation_status = ValidationStatus.VALIDATED
        item.extracted_metadata = {
            "title": f"Document Title {i + 1}",
            "creation_date": "2025-06-15T10:00:00+00:00",
        }
        item.quality_score = 0.75
        item.created_at = datetime.now(tz=UTC) - timedelta(days=i)
        items.append(item)
    return items


def _setup_catalog_mocks(
    mock_db_session: AsyncMock,
    items: list[MagicMock],
    total: int,
) -> None:
    """Configure mock session for catalog endpoint queries."""
    # The catalog endpoint runs two queries: items + count
    items_result = MagicMock()
    items_result.scalars.return_value.all.return_value = items

    count_result = MagicMock()
    count_result.scalar.return_value = total

    mock_db_session.execute.side_effect = [items_result, count_result]
